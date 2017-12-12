"""
This script is intended to flatten an excel file that had a crosstab approach to how respondants answered a question about what informatics functions they are responsible for, or support. There were 32 different functions with one column for responsible and one column for supporting. So 64 different columns.

This script flattens out all of these columns into a single row for each function showing the participation level for the person in the function. 

This should result in 32 rows x 16 respondants + 1 header row or 513 total

Tableau visualization wants row-based data table with each question response a different column.

It would probably take 10 minutes to manually recreate, but instead I wanted to learn and practice python, so this script.

Input is an excel file called analysis.xlsx with a sheet called "Data-Responses-FlatFunctions" that has 70 columns
Person	C/I/O	Role-Level	FunctionLevel	FunctionResp	FunctionSupport	Q4- Asset / property / inventory management	Q4- Center level IT/informatics strategy	Q4- Communication (internal and external presentations, external partners)	Q4- Cross center collaboration	Q4- Data management	Q4- Data use agreements	Q4- Enterprise architecture	Q4- Fact finding / evaluation / alternatives analysis	Q4- HealthIT policy and coordination	Q4- Informatics consultation	Q4- Information security (e.g., manage ISSO)	Q4- IT architecture	Q4- IT contracting	Q4- IT coordination	Q4- IT investment governance (CPIC/EPLC)	Q4- IT project management	Q4- IT services (miscellaneous functions)	Q4- Liaison to CDC business services (ITSO, MISO, etc)	Q4- Liaison to other federal and external agencies (NACCHO, ASTHO, CDC-EU)	Q4- Management of organizational unit / admin code	Q4- Mentoring informatics staff	Q4- Open data	Q4- Representing center in CDC informatics-related bodies (e.g., IRGC, SDP, OAMD, ELC, DCIPHER, platforms, etc.)	Q4- Research & development	Q4- Scientific leadership and management	Q4- Scientific publication/authoring	Q4- SharePoint services	Q4- Solution analysis / business analysis	Q4- Standards	Q4- System/program development and operation (e.g., running VTrckS)	Q4- Technical monitor for cooperative agreement and grants	Q4- Other	Q5- Asset / property / inventory management	Q5- Center level IT/informatics strategy	Q5- Communication (internal and external presentations, external partners)	Q5- Cross center collaboration	Q5- Data management	Q5- Data use agreements	Q5- Enterprise architecture	Q5- Fact finding / evaluation / alternatives analysis	Q5- HealthIT policy and coordination	Q5- Informatics consultation	Q5- Information security (e.g., manage ISSO)	Q5- IT architecture	Q5- IT contracting	Q5- IT coordination	Q5- IT investment governance (CPIC/EPLC)	Q5- IT project management	Q5- IT services (miscellaneous functions)	Q5- Liaison to CDC business services (ITSO, MISO, etc)	Q5- Liaison to other federal and external agencies (NACCHO, ASTHO, CDC-EU)	Q5- Management of organizational unit / admin code	Q5- Mentoring informatics staff	Q5- Open data	Q5- Representing center in CDC informatics-related bodies (e.g., IRGC, SDP, OAMD, ELC, DCIPHER, platforms, etc.)	Q5- Research & development	Q5- Scientific leadership and management	Q5- Scientific publication/authoring	Q5- SharePoint services	Q5- Solution analysis / business analysis	Q5- Standards	Q5- System/program development and operation (e.g., running VTrckS)	Q5- Technical monitor for cooperative agreement and grants	Q5- Other
This is a subset of a full file and has all of the non-relevant columns stripped out for simplicity. There's an example file here, but don't want to have the actual spreadsheet because of name sensitivity. The names aren't relevant to the script anyway

Output is a quote escaped output.csv file with each respondant repeated with 6 columns:
Person	C/I/O	Role-Level	Function	HighestParticipation,	Responsible	Supports
Person=name of responder
C/I/O=center of responder
Role-Level=what kind of position-center, office, advisor
Function=one of the 32 different functions
ParticipationLevel=highest participation level for this person on this function,will be 'R' if highest is responsible (could also support function, check later fields) or 'S' if supporting, or '0' if no participation 
Responsible= either 'R' if the person is resposible for the function, else '0'
Supports=either 'S' if the person supports the function, else '0'
FunctionLevel=if the center has this center, values are "B" if both responsible and supporting, "R" if responsible, "S" if supporting, "0" for none.

Note to future self for lame hacks-only checks rows 2-17 and columns A through BR. If you ever add functions or more than 16 responders, you need to clean this up. Sorry.
"""

import sys
import openpyxl
import csv

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

inputFile = 'analysis.xlsx'
outputFile = 'output.csv'

print 'Start. Working on...' + inputFile

wb = load_workbook(inputFile)

ws = wb.get_sheet_by_name('Data-Responses-FlatFunctions')



numRows = 17 #16 responses plus 1 header

numCols = 70 #only the function responses and a few pre columns

numFunctions = 32 #32 different functions that could be selected for Q4 and Q5

output = open(outputFile,"wb")
writer = csv.writer(output,delimiter=',',quotechar='"',quoting=csv.QUOTE_ALL)
writer.writerow(['Person','C/I/O','Role-Level','Function','Highest Participation','Responsible','Supports'])

processedRows = 0

for row in ws['A2':'BR17']:
	if ws['A'+str(processedRows+2)].value is not None:#only process if the person is named
		for cell in row:
			headerColName = ws[cell.column + '1'].value	
			if headerColName.startswith('Q4'):
				functionName = ws[cell.column + '1'].value[4:]
				responsibleFunction = '0'
				supportsFunction = '0'
				highestFunction = '0'

				#check the equivalent column for supporting
				q4ColNumber = column_index_from_string(cell.column)
				q5ColLetter = get_column_letter(q4ColNumber + numFunctions)
				if ws[q5ColLetter + str(cell.row)].value is not None: #if it has any value, then person is supporting
					supportsFunction = 'S'
					highestFunction = supportsFunction
			
				#check column for responsible
				if cell.value is not None:#if it has any value, then person is resposible
					responsibleFunction ='R'
					highestFunction = responsibleFunction
			
				#print 'functionName= ' + functionName
				#print 'responsibleFunction= ' + responsibleFunction
				#print 'supportsFunction= ' + supportsFunction
				#print 'highestFunctionParticipation= ' + highestFunction

				#name,center,role,function,highest,responsible,suports
				writer.writerow([ws['A' + str(cell.row)].value,ws['B' + str(cell.row)].value,ws['C' + str(cell.row)].value,functionName,highestFunction,responsibleFunction,supportsFunction])
		processedRows +=1

output.close()
print 'Finished. Processed ' + str(processedRows) + ' rows. Check out ' + outputFile
